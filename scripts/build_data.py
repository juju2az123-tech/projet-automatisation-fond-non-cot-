#!/usr/bin/env python3
"""
Construit data/funds_data.js à partir des 2 classeurs Excel Althos :
  - source/Calendriers_de_fonds_Althos.xlsx  (feuilles "Suivi" et "Calendriers")
  - source/Bibliotheque_de_fonds_Althos.xlsx (feuille "Fonds")

Sortie : data/funds_data.js
  window.FUNDS_DATA = { generatedAt, funds: [...], calendar: {isin: {Souscription:[...], Rachat:[...]}} }

À relancer chaque fois que les fichiers sources sont mis à jour (nouveaux calendriers annuels,
nouvelles pénalités, nouveaux fonds). Voir README.md.
"""
import json
import re
import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "source"
OUT = ROOT / "data" / "funds_data.js"

CAL_FILE = SRC / "Calendriers_de_fonds_Althos.xlsx"
BIB_FILE = SRC / "Bibliotheque_de_fonds_Althos.xlsx"


def iso(d):
    if d is None:
        return None
    if isinstance(d, (datetime.datetime, datetime.date)):
        return d.strftime("%Y-%m-%d")
    return None


# ---------------------------------------------------------------------------
# 1) Règles de pénalité de sortie — construites à la main à partir des textes
#    libres trouvés en feuille "Suivi" / "Calendriers par fonds". Chaque règle
#    est volontairement conservatrice : si la formulation ne permet pas de
#    déduire un seuil fiable en mois, on la classe "manuel" et l'app affichera
#    le texte brut sans calcul automatique.
# ---------------------------------------------------------------------------

def parse_penalite(raw):
    """Retourne un dict structuré à partir du texte libre 'Lock-up / pénalité de sortie'."""
    if not raw or not str(raw).strip():
        return {"kind": "inconnue", "raw": None, "tiers": []}

    text = str(raw).strip()
    low = text.lower()

    if "aucune pénalité" in low or "pas de lock-up" in low.replace("lock up", "lock-up"):
        return {"kind": "aucune", "raw": text, "tiers": []}

    # Pénalité dégressive multi-paliers : "0-18 mois 7,5% · 18-36 mois 5% · ..."
    if "dégressiv" in low:
        tiers = []
        for m in re.finditer(r"(\d+)\s*-\s*(\d+)\s*mois\s*(\d+(?:,\d+)?)\s*%", text):
            lo, hi, rate = m.groups()
            tiers.append({
                "minMonths": int(lo),
                "maxMonths": int(hi),
                "rate": float(rate.replace(",", ".")),
            })
        # dernier palier "au-delà de 5 ans 0%"
        m2 = re.search(r"au-del[àa] de (\d+)\s*ans?\s*(\d+(?:,\d+)?)\s*%", low)
        if m2:
            tiers.append({
                "minMonths": int(m2.group(1)) * 12,
                "maxMonths": None,
                "rate": float(m2.group(2).replace(",", ".")),
            })
        return {"kind": "degressif", "raw": text, "tiers": tiers}

    # Plusieurs taux distincts mentionnés (ex : parts différentes avec pénalités
    # différentes) : trop ambigu pour un calcul automatique fiable -> manuel.
    all_rates = {r.replace(",", ".") for r in re.findall(r"(\d+(?:,\d+)?)\s*%", text)}
    if len(all_rates) > 1:
        return {"kind": "manuel", "raw": text, "tiers": []}

    # Cas à seuil simple : "X% ... avant/dans les N mois|ans de détention|souscription|émission"
    m = re.search(r"(\d+(?:,\d+)?)\s*%", text)
    m_months = re.search(r"(\d+)\s*(?:premiers?\s+|premi[eè]res?\s+)?(mois|ans?)\b", low)
    if m and m_months:
        n = int(m_months.group(1))
        unit = m_months.group(2)
        months = n * 12 if unit.startswith("an") else n
        soft = "soft lock-up" in low or "pas de hard lock-up" in low
        return {
            "kind": "soft" if soft else "seuil",
            "raw": text,
            "tiers": [{"minMonths": 0, "maxMonths": months, "rate": float(m.group(1).replace(",", "."))}],
        }

    # Formulation ambiguë (ex : "Pénalité de 4% en cours (au 25/09/2024)") ou
    # multi-parts (ex : Hg Fusion "5% pendant 12 mois (part F2) ; 2% ... ancienne part")
    return {"kind": "manuel", "raw": text, "tiers": []}


# ---------------------------------------------------------------------------
# 2) Lecture feuille "Suivi" (liste maîtresse des fonds non cotés suivis)
# ---------------------------------------------------------------------------

def load_suivi(wb):
    ws = wb["Suivi"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    funds = {}
    for r in rows:
        isin, nom, cat, souscat = r[0], r[1], r[2], r[3]
        calendrier_recu = r[4]
        couverture_jusquau = r[9]
        statut_couverture = r[10]
        penalite_raw = r[11]
        if not nom:
            continue
        key = isin or ("NOISIN::" + nom)
        funds[key] = {
            "isin": isin,
            "nom": nom.strip() if isinstance(nom, str) else nom,
            "categorie": cat,
            "sousCategorie": souscat,
            "calendrierRecu": calendrier_recu,
            "couvertureJusquau": iso(couverture_jusquau),
            "statutCouverture": statut_couverture,
            "penalite": parse_penalite(penalite_raw),
        }
    return funds


# ---------------------------------------------------------------------------
# 3) Fonds supplémentaires présents dans "Calendriers par fonds" mais absents
#    de "Suivi" (ex : Antheor Private Infrastructure) → on les ajoute.
# ---------------------------------------------------------------------------

def merge_extra_from_calendriers_par_fonds(wb, funds):
    ws = wb["Calendriers par fonds"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for r in rows:
        isin, nom, cat, souscat = r[0], r[1], r[2], r[3]
        penalite_raw = r[15]
        if not isin or not nom:
            continue
        if not re.match(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$|^[A-Z0-9]{6,15}$", str(isin)):
            # ligne "Mois" ou autre sous-en-tête de bloc, pas une vraie ligne fonds
            continue
        if isin not in funds:
            funds[isin] = {
                "isin": isin,
                "nom": nom.strip() if isinstance(nom, str) else nom,
                "categorie": cat,
                "sousCategorie": souscat,
                "calendrierRecu": None,
                "couvertureJusquau": None,
                "statutCouverture": None,
                "penalite": parse_penalite(penalite_raw),
            }
        elif funds[isin]["penalite"]["kind"] == "inconnue" and penalite_raw:
            funds[isin]["penalite"] = parse_penalite(penalite_raw)
    return funds


# ---------------------------------------------------------------------------
# 4) Feuille "Calendriers" (une ligne = fonds/année/type/mois) → calendrier
#    mensuel exploitable par l'app pour calculer les prochaines dates d'ordre.
# ---------------------------------------------------------------------------

def load_calendrier(wb):
    ws = wb["Calendriers"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    calendar = {}
    for r in rows:
        (cle, isin, nom, societe, annee, type_, periode, cutoff, valorisation,
         execution, publication_vl, reglement_cash, note, source, statut,
         fichier, horodatage, passee_avenir) = r[:18]
        if not isin or not type_:
            continue
        entry = {
            "periode": periode,
            "cutoff": iso(cutoff),
            "valorisation": iso(valorisation),
            "execution": iso(execution),
            "publicationVL": iso(publication_vl),
            "reglementCash": iso(reglement_cash),
            "statut": statut,
        }
        calendar.setdefault(isin, {}).setdefault(type_, []).append(entry)

    # tri chronologique par date d'exécution (fallback cutoff) pour chaque liste
    def sort_key(e):
        return e["execution"] or e["cutoff"] or e["valorisation"] or "9999-99-99"

    for isin, by_type in calendar.items():
        for type_, entries in by_type.items():
            entries.sort(key=sort_key)
    return calendar


MOIS_RE = re.compile(
    r"^(Janvier|F[ée]vrier|Mars|Avril|Mai|Juin|Juillet|Ao[uû]t|Septembre|Octobre|Novembre|D[ée]cembre)"
    r"\s+\d{4}$"
)


def load_calendrier_par_fonds(wb):
    """
    La feuille "Calendriers par fonds" liste tout l'univers de fonds (comme "Fonds"), et pour
    certains d'entre eux seulement, insère juste en dessous un mini-tableau calendrier (ligne
    d'en-tête "Mois", puis une ligne par mois : cut-off / VL / exécution / publication / cash,
    pour la souscription ET le rachat côte à côte). Ce sont "les lignes du dessous" : pour
    certains fonds (ex. Schroder Semi Liquid GPE), c'est l'UNIQUE source de calendrier — ils
    n'apparaissent pas du tout dans la feuille "Calendriers" à plat.
    """
    ws = wb["Calendriers par fonds"]
    max_row = ws.max_row
    calendar = {}

    for r in range(2, max_row + 1):
        if ws.cell(row=r, column=1).value != "Mois":
            continue
        # Le fonds concerné est la ligne non vide la plus proche au-dessus (ISIN en colonne A).
        isin = None
        for r2 in range(r - 1, max(r - 8, 0), -1):
            v = ws.cell(row=r2, column=1).value
            if v:
                isin = v
                break
        if not isin or not re.match(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$|^[A-Z0-9]{6,15}$", str(isin)):
            continue

        r2 = r + 1
        while r2 <= max_row:
            periode = ws.cell(row=r2, column=1).value
            if not periode or not MOIS_RE.match(str(periode).strip()):
                break
            vals = [ws.cell(row=r2, column=c).value for c in range(2, 12)]
            (sous_cutoff, sous_val, sous_exec, sous_pub, sous_cash,
             rac_cutoff, rac_val, rac_exec, rac_pub, rac_cash) = vals

            # Repli sur la date d'exécution quand la VL n'est pas publiée séparément dans ce
            # tableau, pour ne jamais afficher une date de valorisation vide (=> "00/01/1900"
            # une fois passée dans TEXT() côté formule Excel).
            if sous_cutoff or sous_exec:
                calendar.setdefault(isin, {}).setdefault("Souscription", []).append({
                    "periode": periode,
                    "cutoff": iso(sous_cutoff), "valorisation": iso(sous_val or sous_exec),
                    "execution": iso(sous_exec), "publicationVL": iso(sous_pub),
                    "reglementCash": iso(sous_cash), "statut": "Calendriers par fonds",
                })
            if rac_cutoff or rac_exec:
                calendar.setdefault(isin, {}).setdefault("Rachat", []).append({
                    "periode": periode,
                    "cutoff": iso(rac_cutoff), "valorisation": iso(rac_val or rac_exec),
                    "execution": iso(rac_exec), "publicationVL": iso(rac_pub),
                    "reglementCash": iso(rac_cash), "statut": "Calendriers par fonds",
                })
            r2 += 1

    return calendar


def merge_calendriers(primary, extra):
    """Fusionne `extra` dans `primary`, sans dupliquer une échéance déjà connue (même couple
    isin/type/cutoff). `primary` (feuille "Calendriers", entretenue à la main) reste la source
    de vérité en cas de chevauchement ; `extra` ne fait que combler les trous — nouveaux fonds
    (ex. Schroder Semi Liquid GPE) ou mois absents de `primary` pour un fonds déjà connu."""
    for isin, by_type in extra.items():
        for type_, entries in by_type.items():
            existing = primary.setdefault(isin, {}).setdefault(type_, [])
            known_cutoffs = {e["cutoff"] for e in existing}
            for e in entries:
                if e["cutoff"] not in known_cutoffs:
                    existing.append(e)
                    known_cutoffs.add(e["cutoff"])

    def sort_key(e):
        return e["execution"] or e["cutoff"] or e["valorisation"] or "9999-99-99"

    for isin, by_type in primary.items():
        for type_, entries in by_type.items():
            entries.sort(key=sort_key)
    return primary


# ---------------------------------------------------------------------------
# 5) Feuille "Fonds" de la Bibliothèque → attributs complémentaires (devise,
#    SRI, frais, liquidité, temporalité, lien DICI) rattachés par ISIN.
# ---------------------------------------------------------------------------

def load_bibliotheque(wb):
    ws = wb["Fonds"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    attrs = {}
    for r in rows:
        isin, nom, cat, souscat, devise, sri, frais, surperf, dici = r[0:9]
        temporalite, statut, maj, commentaire, liquidite, sfdr = r[16], r[17], r[18], r[19], r[20], r[21]
        if not isin or not nom:
            continue
        attrs[isin] = {
            "devise": devise,
            "sri": sri,
            "fraisCourants": frais,
            "commissionSurperformance": surperf,
            "lienDici": dici,
            "temporalite": temporalite,
            "liquidite": liquidite,
            "sfdr": sfdr,
        }
    return attrs


def main():
    wb_cal = openpyxl.load_workbook(CAL_FILE, data_only=True)
    wb_bib = openpyxl.load_workbook(BIB_FILE, data_only=True)

    funds = load_suivi(wb_cal)
    funds = merge_extra_from_calendriers_par_fonds(wb_cal, funds)
    calendar = load_calendrier(wb_cal)
    calendar = merge_calendriers(calendar, load_calendrier_par_fonds(wb_cal))
    bib_attrs = load_bibliotheque(wb_bib)

    fund_list = []
    for key, f in funds.items():
        isin = f["isin"]
        f["key"] = key  # identifiant stable côté app (ISIN, ou "NOISIN::Nom" si pas d'ISIN)
        f["hasCalendar"] = bool(isin and isin in calendar)
        if isin and isin in bib_attrs:
            f.update({k: v for k, v in bib_attrs[isin].items()})
        fund_list.append(f)

    # Tri alpha par nom pour un menu déroulant lisible
    fund_list.sort(key=lambda f: (f["nom"] or ""))

    data = {
        "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "sourceFiles": [CAL_FILE.name, BIB_FILE.name],
        "funds": fund_list,
        "calendar": calendar,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write("// Fichier généré automatiquement par scripts/build_data.py — ne pas éditer à la main.\n")
        fh.write("window.FUNDS_DATA = ")
        json.dump(data, fh, ensure_ascii=False, indent=2)
        fh.write(";\n")

    print(f"OK — {len(fund_list)} fonds, {len(calendar)} calendriers -> {OUT}")
    kinds = {}
    for f in fund_list:
        k = f["penalite"]["kind"]
        kinds[k] = kinds.get(k, 0) + 1
    print("Répartition des règles de pénalité :", kinds)


if __name__ == "__main__":
    main()
