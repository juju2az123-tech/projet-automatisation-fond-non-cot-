#!/usr/bin/env python3
"""
Ajoute au classeur de consolidation client (ex : source/ConsolidationTemplateAlthosAI_V5.xlsx)
une nouvelle feuille "Calendrier de sortie" — un tableau COMPACT listant uniquement les fonds
que le client détient réellement (montant non nul dans sa consolidation) ET pour lesquels un
calendrier de RACHAT (sortie) officiel est connu dans la base Althos. Ni les fonds cotés, ni les
fonds non détenus, ni les fonds sans calendrier de rachat n'apparaissent : ce n'est pas une copie
de la feuille "Consolidation", juste la liste utile pour répondre à un client qui demande ses
délais de sortie. Pour chaque fonds retenu :

  - Titulaire (Monsieur / Madame / société... — vide si le fichier n'a qu'un seul titulaire)
  - Fonds, ISIN
  - Date d'investissement (à saisir par le conseiller)
  - Rachat — ordre avant / VL / exécuté / publié / cash reçu (valeurs reprises telles quelles
    de la base, pour la prochaine échéance de rachat à partir d'aujourd'hui)
  - Pénalité de sortie (vide si le délai de pénalité est dépassé ; message dans tous les autres
    cas, calculé par rapport à la date d'investissement saisie)

Présentation reprise de la feuille Consolidation elle-même (même police, mêmes couleurs) :
bandeaux de catégorie (beige, gras) au-dessus des fonds qu'ils regroupent, en-tête de tableau
dans la même couleur que celui de Consolidation. Un même fonds détenu par plusieurs titulaires
(ex. Monsieur ET Madame, chacun via son propre contrat) donne une ligne par titulaire.

Toutes les colonnes calculées sont des FORMULES Excel (recalculées à chaque ouverture, à la
date du jour), et s'appuient sur deux feuilles de données ajoutées et masquées :
  - BDD_Calendrier   (calendrier mensuel des fonds, dérivé de Calendriers_de_fonds_Althos.xlsx)
  - BDD_Penalites    (règles de pénalité de sortie structurées, même parseur que build_data.py)

Usage :
    python3 scripts/build_client_workbook.py [chemin_consolidation.xlsx] [chemin_sortie.xlsx] [date_retrait_simulee]

Par défaut lit source/ConsolidationTemplateAlthosAI_V5.xlsx et écrit
output/ConsolidationTemplateAlthosAI_V5_avec_calendrier.xlsx

Le 3e argument optionnel (format AAAA-MM-JJ, ex. 2026-12-09) simule une date de retrait future :
tout le tableau (échéances de rachat, durée de détention, pénalité de sortie) est alors calculé
comme si on s'y trouvait déjà, plutôt qu'à la date du jour.
"""
import sys
import re
import datetime
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_data import (  # noqa: E402
    load_suivi, merge_extra_from_calendriers_par_fonds, load_calendrier,
    load_calendrier_par_fonds, merge_calendriers, attach_duree_vie,
)

CAL_FILE = ROOT / "source" / "Calendriers_de_fonds_Althos.xlsx"



# ---------------------------------------------------------------------------
# Détection dynamique de la structure de la feuille "Consolidation" (aucun
# numéro de ligne/colonne n'est supposé fixe, pour s'adapter à un fichier
# client réel dont la mise en page peut différer du template).
# ---------------------------------------------------------------------------

def find_header_row(ws):
    for r in range(1, 16):
        if (ws.cell(row=r, column=1).value or "").strip() == "Support":
            return r
    return None


def build_grid_border(ws, header_row):
    """Couleur ET épaisseur de quadrillage reprises telles quelles (référence de thème + teinte,
    PAS une valeur fixe) de la bordure déjà utilisée par l'en-tête de Consolidation — sa couleur
    réelle dépend du thème propre à chaque classeur (beige dans certains, bleu dans d'autres).
    header_border.top est le cadre extérieur ÉPAIS de l'en-tête (volontairement plus marqué),
    jamais représentatif du quadrillage interne fin utilisé partout ailleurs dans le tableau ;
    header_border.left, elle, EST cette bordure fine — c'est elle qu'il faut reprendre en
    priorité, sans forcer une épaisseur différente. Renvoie une bordure complète (4 côtés) prête à
    appliquer à une cellule."""
    header_border = ws.cell(row=header_row, column=1).border
    ref_side = header_border.left or header_border.right or header_border.bottom or header_border.top
    side = (copy(ref_side) if ref_side is not None and ref_side.style
            else Side(style="thin", color="FFDDCCB8"))
    return Border(top=side, left=side, bottom=side, right=side)


PENALTY_PREFIXES = {
    "ferme": "Fonds fermé : aucun rachat possible.",
    "manuel": "À VÉRIFIER MANUELLEMENT : ",
    "concerne1": "Pénalité de sortie : ",
    "concerne2": "Durée de détention actuelle : XXX mois.",
}


def estimate_penalty_lines(kind, raw_len, duree_len, chars_per_line):
    """Nombre de lignes à prévoir pour le texte de pénalité d'un fonds donné, une fois la mise en
    forme finale appliquée (2 phrases séparées par un retour à la ligne explicite pour "ferme" et
    les pénalités actives datées) — calculé à partir de la longueur RÉELLE du texte de la base
    (connue à la génération), pas devinée : un texte source inhabituellement long (ça arrive)
    donne une ligne plus haute plutôt qu'un texte coupé."""
    def lines_for(length):
        return max(1, -(-length // chars_per_line))  # ceil

    if kind == "ferme":
        lines = lines_for(len(PENALTY_PREFIXES["ferme"]))
        if duree_len:
            lines += lines_for(duree_len)
        return lines
    if kind == "manuel":
        return lines_for(len(PENALTY_PREFIXES["manuel"]) + raw_len)
    if kind in ("seuil", "soft", "degressif"):
        return lines_for(len(PENALTY_PREFIXES["concerne1"]) + raw_len) + lines_for(len(PENALTY_PREFIXES["concerne2"]))
    return 1  # aucune / inconnue / date manquante / date future : message court


def height_for_lines(lines, font_size):
    """Hauteur de ligne (en points) pour `lines` lignes de texte à la taille de police
    `font_size` — approximation généreuse du ratio hauteur de ligne / taille de police (~1.5,
    contre ~1.2 pour une police système classique) : la police réelle utilisée dans ces
    classeurs (Montserrat) est plus large et rend visuellement plus haute qu'une police par
    défaut, donc la marge est volontairement large pour ne jamais couper le texte de justesse."""
    return round(lines * (font_size * 1.6 + 4))


# Volontairement plus haut qu'un minimum "juste" : sans moteur de rendu disponible pour vérifier
# pixel par pixel, on préfère une ligne visiblement plus haute que nécessaire à un texte coupé
# dans le classeur final.
DEFAULT_PENALTY_ROW_HEIGHT = 75


def find_total_column(ws, header_row):
    for c in range(3, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and str(v).strip().upper() == "TOTAL":
            return c
    return None


def find_last_header_column(ws, header_row):
    """Dernière colonne non vide de la ligne d'en-tête de Consolidation, pour savoir où ajouter
    de nouvelles colonnes sans écraser celles qui existent déjà."""
    last = 1
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value not in (None, ""):
            last = c
    return last


def is_merged_follower(ws, row, col):
    """Vrai si la cellule est "esclave" d'une fusion (pas la cellule maîtresse en haut à
    gauche) : openpyxl refuse toute écriture dessus (AttributeError)."""
    from openpyxl.cell.cell import MergedCell
    return isinstance(ws.cell(row=row, column=col), MergedCell)


def next_safe_column(ws, row, start):
    """Première colonne à partir de `start` qui n'est pas une cellule esclave d'une fusion sur
    cette ligne — pour ne jamais faire atterrir une nouvelle colonne au milieu d'une fusion
    existante (ex. "Mouvements en cours" fusionnée sur 2 colonnes dans certains fichiers)."""
    c = start
    while is_merged_follower(ws, row, c):
        c += 1
    return c


def find_column_by_header_text(ws, header_row, text):
    """Colonne du premier en-tête dont le texte correspond exactement (après trim) à `text`, ou
    None si absent — pour repérer une colonne "connue" comme "Mouvements en cours" par son
    intitulé plutôt que par une position codée en dur."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip() == text:
            return c
    return None


def merged_block_end_column(ws, row, start_col):
    """Dernière colonne du bloc fusionné (horizontalement) démarrant à `start_col` sur `row`."""
    end = start_col
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= start_col <= mc.max_col:
            end = max(end, mc.max_col)
    return end


def shift_column_block(ws, from_start, from_end, to_start):
    """Déplace tout un bloc de colonnes [from_start..from_end] vers [to_start..], sur toutes les
    lignes de la feuille : valeurs, styles, formats et fusions (jamais les formules elles-mêmes —
    aucune bibliothèque ExcelJS/openpyxl ne réécrit les références de colonnes lors d'un
    déplacement, mais sans risque ici puisque ce bloc, tel qu'observé sur les fichiers réels, ne
    contient que des valeurs statiques, jamais de formule). Utilisé pour repousser une colonne
    existante (ex. "Mouvements en cours") et faire de la place aux 2 nouvelles colonnes."""
    if to_start == from_start:
        return
    width = from_end - from_start
    max_row = ws.max_row
    shift = to_start - from_start

    moved_merges = [mc for mc in list(ws.merged_cells.ranges)
                     if mc.min_col >= from_start and mc.max_col <= from_end]
    for mc in moved_merges:
        ws.unmerge_cells(str(mc))

    for r in range(1, max_row + 1):
        for i in range(width + 1):
            src = ws.cell(row=r, column=from_start + i)
            val = src.value
            style = copy(src._style)
            num_fmt = src.number_format
            dst = ws.cell(row=r, column=to_start + i)
            dst.value = val
            dst._style = style
            dst.number_format = num_fmt
            ws._cells.pop((r, from_start + i), None)  # remet la cellule d'origine à l'état vierge

    for i in range(width + 1):
        old_letter = get_column_letter(from_start + i)
        new_letter = get_column_letter(to_start + i)
        if old_letter in ws.column_dimensions and ws.column_dimensions[old_letter].width:
            ws.column_dimensions[new_letter].width = ws.column_dimensions[old_letter].width
            del ws.column_dimensions[old_letter]

    for mc in moved_merges:
        ws.merge_cells(start_row=mc.min_row, start_column=mc.min_col + shift,
                        end_row=mc.max_row, end_column=mc.max_col + shift)


def extend_print_area(ws, min_col_needed):
    """Étend la zone d'impression déjà définie sur `ws` (le contour bleu visible dans Excel) pour
    couvrir au minimum jusqu'à la colonne `min_col_needed` — sans jamais la réduire. Ne fait rien
    si aucune zone d'impression n'est définie sur cette feuille."""
    area = ws.print_area
    if not area:
        return
    range_part = area.split("!")[-1]  # ws.print_area peut être préfixé par le nom de la feuille
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(range_part)
    if min_col_needed > max_col:
        ws.print_area = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(min_col_needed)}{max_row}"


def extend_merge_right(ws, row, min_col_needed):
    """Étend vers la droite la plage fusionnée qui commence à (row, 1) — typiquement le bandeau
    de titre de Consolidation — pour couvrir au minimum jusqu'à `min_col_needed`, sans jamais la
    réduire. Ne fait rien si cette cellule n'est pas fusionnée."""
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row == row and mc.min_col == 1:
            if min_col_needed > mc.max_col:
                ws.unmerge_cells(str(mc))
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min_col_needed)
            return


def has_two_row_header(ws, header_row, col):
    """Vrai si l'en-tête de cette colonne occupe 2 lignes fusionnées (header_row:header_row+1),
    comme "Support"/"Dernière Valeur Liquidative" dans les fichiers Althos — pour aligner les 2
    nouvelles colonnes sur le même bleu marine "en hauteur" que leurs voisines."""
    from openpyxl.cell.cell import MergedCell
    if isinstance(ws.cell(row=header_row, column=col), MergedCell):
        return False
    if not isinstance(ws.cell(row=header_row + 1, column=col), MergedCell):
        return False
    for mc in ws.merged_cells.ranges:
        if mc.min_row == header_row and mc.max_row == header_row + 1 and mc.min_col <= col <= mc.max_col:
            return True
    return False


def _is_sum_total(v):
    """Vrai si `v` (valeur d'une cellule TOTAL) correspond à une vraie agrégation de fonds —
    =SUM(...) sur ses colonnes de contrat, ou un nombre déjà calculé — plutôt qu'une simple
    formule de renvoi vers une autre cellule (=A8, =IFERROR(U312,"/")...), utilisée par certains
    petits tableaux récapitulatifs hors-tableau plus bas sur la feuille ("dont actions
    européennes : 15 %"...) qui ne sont PAS des lignes de fonds malgré un texte en colonne A."""
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        return v.strip().upper().lstrip("=").startswith("SUM(")
    return False


def classify_rows(ws, header_row, total_col=None):
    """Une ligne de catégorie (bandeau) a un fond uni et pas d'ISIN en colonne B — jamais une
    vraie ligne de fonds. Le gras n'est PAS un critère fiable : certains fichiers clients ne
    mettent en gras que les catégories de premier niveau, pas les sous-catégories (ex. "dont
    actions européennes"), qui restent pourtant de vrais bandeaux (fond uni, pas d'ISIN) à traiter
    de la même façon.

    Associe aussi à chaque ligne fonds le libellé de la dernière catégorie rencontrée
    au-dessus d'elle (row_to_category), pour reproduire les mêmes bandeaux de catégorie
    dans la feuille générée.
    """
    cat_rows, fund_rows = [], []
    row_to_category = {}
    # Fond (beige clair ou foncé selon le niveau d'imbrication de LA catégorie) de la ligne de
    # catégorie la plus proche au-dessus de chaque fonds — pour pouvoir reproduire fidèlement la
    # même teinte ailleurs (Consolidation utilise 2 teintes de beige distinctes selon le niveau).
    row_to_category_fill = {}
    current_category = None
    current_category_fill = None
    for r in range(header_row + 2, ws.max_row + 1):  # header_row+1 = ligne de total général
        a = ws.cell(row=r, column=1)
        b = ws.cell(row=r, column=2)
        if a.value is None:
            continue
        if a.fill.patternType == "solid" and b.value is None:
            cat_rows.append(r)
            current_category = str(a.value).strip()
            current_category_fill = copy(a.fill)
        else:
            if total_col and not _is_sum_total(ws.cell(row=r, column=total_col).value):
                continue
            fund_rows.append(r)
            row_to_category[r] = current_category
            row_to_category_fill[r] = current_category_fill
    return cat_rows, fund_rows, row_to_category, row_to_category_fill


def holding_amount(ws, row, total_col):
    """Somme des colonnes 'Contrat' (3..total_col-1) ; None si aucune colonne TOTAL trouvée."""
    if not total_col:
        return None
    total = 0
    for c in range(3, total_col):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, (int, float)):
            total += v
    return total


def detect_owner_labels(ws, header_row, total_col):
    """Libellé du titulaire (Monsieur / Madame / société / prénom...) par colonne 'Contrat',
    lu sur la ligne juste au-dessus de l'en-tête (souvent fusionnée sur plusieurs colonnes
    contrat). openpyxl ne reporte la valeur que sur la cellule ancre d'une fusion : on résout
    donc explicitement les plages fusionnées qui couvrent cette ligne. Absent ou vide pour un
    client à titulaire unique (pas de subdivision) : toutes les colonnes retombent alors sur "".
    """
    labels = {}
    owner_row = header_row - 1
    if not total_col or owner_row < 1:
        return labels
    merged_lookup = {}
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= owner_row <= mc.max_row:
            anchor_val = ws.cell(row=mc.min_row, column=mc.min_col).value
            for c in range(mc.min_col, mc.max_col + 1):
                merged_lookup[c] = anchor_val
    for c in range(3, total_col):
        v = merged_lookup.get(c, ws.cell(row=owner_row, column=c).value)
        labels[c] = str(v).strip() if v is not None else ""
    return labels


def holding_by_owner(ws, row, total_col, owner_labels):
    """Répartit le montant détenu d'une ligne fonds par titulaire (clé "" si non subdivisé)."""
    if not total_col:
        return [("", None)]  # structure inconnue : une seule ligne, montant indéterminé
    amounts = {}
    for c in range(3, total_col):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, (int, float)):
            label = owner_labels.get(c, "")
            amounts[label] = amounts.get(label, 0) + v
    return list(amounts.items())


def extract_owner_civilities(title_text):
    """Extrait {"nom complet (dans l'ordre du titre)": "M."|"Mme"} depuis le texte du titre de
    Consolidation (ex. "Consolidation Monsieur Gérard LENO**** et Madame Marie MAUN***"), pour
    harmoniser les bandeaux de titulaire de "Calendrier de sortie" uniquement (jamais
    Consolidation elle-même, qui garde son propre texte tel quel). Ne devine jamais un genre : ne
    renvoie que ce que le titre énonce explicitement."""
    out = {}
    if not title_text:
        return out
    for m in re.finditer(r"(Monsieur|Madame)\s+([^,;]+?)(?=\s+(?:et|Monsieur|Madame)\b|$)", str(title_text), re.IGNORECASE):
        civ = "M." if m.group(1).lower() == "monsieur" else "Mme"
        out[m.group(2).strip()] = civ
    return out


def harmonize_owner_label(label, civilities):
    """Reformate un libellé de titulaire "NOM Prénom" (tel que lu dans Consolidation) en
    "M. Prénom NOM" / "Mme Prénom NOM" quand le titre de Consolidation permet une correspondance
    sans ambiguïté (mêmes mots, sans deviner un genre) ; renvoie le libellé d'origine inchangé
    sinon (ex. une société)."""
    if not label:
        return label
    label_tokens = "|".join(sorted(t.lower() for t in label.split()))
    for name, civ in civilities.items():
        name_tokens = "|".join(sorted(t.lower() for t in name.split()))
        if label_tokens and label_tokens == name_tokens:
            return f"{civ} {name}"
    return label


# ---------------------------------------------------------------------------
# Pénalité -> 9 valeurs numériques exploitables par une formule IFS en cascade
# (4 paliers "seuil" + 1 taux "au-delà"), quel que soit le nombre réel de
# paliers d'origine (seuil simple, dégressif à N paliers, etc.)
# ---------------------------------------------------------------------------

def build_tier_columns(pen):
    kind = pen["kind"]
    if kind in ("seuil", "soft"):
        t = pen["tiers"][0]
        display = [{"max": t["maxMonths"], "rate": t["rate"]}, {"max": None, "rate": 0}]
    elif kind == "degressif":
        display = [{"max": t["maxMonths"], "rate": t["rate"]} for t in pen["tiers"]]
    else:
        return [0, 0, 0, 0, 0, 0, 0, 0, 0]

    conditions = display[:-1]
    tail_rate = display[-1]["rate"]
    while len(conditions) < 4:
        last_max = conditions[-1]["max"] if conditions else 0
        conditions.append({"max": last_max, "rate": tail_rate})
    conditions = conditions[:4]

    out = []
    for c in conditions:
        out.append(c["max"])
        out.append(c["rate"])
    out.append(tail_rate)
    return out  # [Max1,Rate1,Max2,Rate2,Max3,Rate3,Max4,Rate4,Rate5]


# ---------------------------------------------------------------------------
# BDD_Calendrier
# ---------------------------------------------------------------------------

def write_bdd_calendrier(wb, calendar):
    ws = wb.create_sheet("BDD_Calendrier")
    headers = ["ISIN", "Nom", "Type", "Cutoff", "Valorisation", "Execution",
               "PublicationVL", "ReglementCash", "CleComposite"]
    ws.append(headers)
    row_i = 2
    for isin, by_type in calendar.items():
        for type_, entries in by_type.items():
            expanded_types = ["Souscription", "Rachat"] if type_ == "Souscription et rachat" else [type_]
            for etype in expanded_types:
                for e in entries:
                    cutoff = e["cutoff"]
                    key = f"{isin}|{etype}|{cutoff}" if cutoff else ""
                    ws.append([
                        isin, "", etype,
                        _to_date(e["cutoff"]), _to_date(e["valorisation"]), _to_date(e["execution"]),
                        _to_date(e["publicationVL"]), _to_date(e["reglementCash"]), key,
                    ])
                    row_i += 1
    for col, width in zip("ABCDEFGHI", [14, 4, 12, 12, 12, 12, 12, 12, 26]):
        ws.column_dimensions[col].width = width
    for c in ws[1]:
        c.font = openpyxl.styles.Font(bold=True)
    ws.sheet_state = "hidden"
    return ws


def _to_date(iso):
    if not iso:
        return None
    y, m, d = iso.split("-")
    return datetime.date(int(y), int(m), int(d))


# ---------------------------------------------------------------------------
# BDD_Penalites
# ---------------------------------------------------------------------------

def write_bdd_penalites(wb, funds):
    ws = wb.create_sheet("BDD_Penalites")
    headers = ["ISIN", "Nom", "Kind", "Max1", "Rate1", "Max2", "Rate2",
               "Max3", "Rate3", "Max4", "Rate4", "Rate5", "RawText", "DureeVie"]
    ws.append(headers)
    for key, f in funds.items():
        isin = f["isin"]
        if not isin:
            continue
        pen = f["penalite"]
        tiers9 = build_tier_columns(pen)
        ws.append([isin, f["nom"], pen["kind"], *tiers9, pen["raw"] or "", pen.get("dureeVie") or ""])
    for col, width in zip("ABCDEFGHIJKLMN", [14, 34, 11, 7, 7, 7, 7, 7, 7, 7, 7, 7, 60, 60]):
        ws.column_dimensions[col].width = width
    for c in ws[1]:
        c.font = openpyxl.styles.Font(bold=True)
    ws.sheet_state = "hidden"
    return ws


# ---------------------------------------------------------------------------
# Nouvelle feuille visible "Calendrier de sortie"
# ---------------------------------------------------------------------------

def format_date_fr(d):
    """« 18/08/2026 », au format numérique jour/mois/année."""
    return d.strftime("%d/%m/%Y")


HELPER_NAMES = [
    "has_sortie", "next_cutoff_sortie", "next_val_sortie", "next_exec_sortie",
    "next_pub_sortie", "next_cash_sortie",
    "months_held", "pen_found", "kind", "raw", "duree_vie",
    "max1", "rate1", "max2", "rate2", "max3", "rate3", "max4", "rate4", "rate5",
    "rate_now", "ref_date",
]
HELPER_FIRST_COL = 11  # K (colonnes visibles jusqu'en I : Fonds, ISIN, Date, 5 dates de rachat, Pénalité)


def helper_col(name):
    idx = HELPER_FIRST_COL + HELPER_NAMES.index(name)
    return get_column_letter(idx)


def has_rachat_calendar(calendar, isin):
    """Un fonds n'a de calendrier "de sortie" que s'il a des échéances de type Rachat."""
    by_type = calendar.get(isin) if isin else None
    return bool(by_type and (by_type.get("Rachat") or by_type.get("Souscription et rachat")))


def build_exit_sheet(wb, src_ws, calendar, cal_last_row, pen_last_row, funds_by_isin, ref_date_iso=None):
    def cal(col):
        return f"BDD_Calendrier!${col}$2:${col}${cal_last_row}"

    def pen(col):
        return f"BDD_Penalites!${col}$2:${col}${pen_last_row}"

    # Date de référence utilisée à la place de TODAY() dans tous les calculs (durée de détention,
    # pénalité active, prochaine échéance de rachat) : soit la date du jour (par défaut, formule
    # toujours à jour à l'ouverture), soit une date de retrait hypothétique choisie par le
    # conseiller pour simuler "et si le client sortait à telle date future ?". Toujours dans une
    # cellule dédiée (jamais TODAY() en dur dans les formules) pour que TOUT le tableau — dates de
    # rachat comprises — se recalcule par rapport à cette même date de référence.
    RD = helper_col("ref_date")
    # Si le conseiller a choisi (via le calendrier) la date du jour elle-même, ce n'est pas une
    # simulation — le comportement doit rester strictement identique à un champ laissé vide
    # (formule =TODAY() toujours à jour, pas de bandeau), plutôt que de figer une date qui
    # deviendrait fausse dès le lendemain.
    is_simulating = bool(ref_date_iso) and ref_date_iso != datetime.date.today().isoformat()

    ws = wb.create_sheet("Calendrier de sortie")
    wb._sheets.remove(ws)
    wb._sheets.insert(wb._sheets.index(src_ws) + 1, ws)  # right after "Consolidation"

    if is_simulating:
        y, m, d = (int(p) for p in ref_date_iso.split("-"))
        ref_cell = ws[f"{RD}1"]
        ref_cell.value = datetime.datetime(y, m, d)
        ref_cell.number_format = "dd/mm/yyyy"
    else:
        ws[f"{RD}1"] = "=TODAY()"

    header_row = find_header_row(src_ws)
    if header_row is None:
        raise ValueError('En-tête "Support" introuvable dans les 15 premières lignes de la feuille Consolidation.')
    total_col = find_total_column(src_ws, header_row)
    cat_rows, fund_rows, row_to_category, row_to_category_fill = classify_rows(src_ws, header_row, total_col)
    owner_labels = detect_owner_labels(src_ws, header_row, total_col)

    # Fonds détenus (répartis par titulaire de contrat) ET dotés d'un calendrier de RACHAT connu.
    # Un même fonds détenu par plusieurs titulaires (ex. Monsieur ET Madame, chacun via son
    # propre contrat) donne une ligne par titulaire, pour une date d'investissement et un statut
    # de pénalité propres à chacun.
    selected = []
    for r in fund_rows:
        isin_raw = src_ws.cell(row=r, column=2).value
        isin = isin_raw.strip() if isinstance(isin_raw, str) else isin_raw
        fund = funds_by_isin.get(isin) if isin else None
        # Un fonds fermé (aucune sortie possible hors cas exceptionnel) est toujours retenu même
        # sans calendrier de rachat : c'est justement l'information à signaler au conseiller.
        is_ferme = bool(fund and fund.get("penalite", {}).get("kind") == "ferme")
        if not fund or not (is_ferme or has_rachat_calendar(calendar, isin)):
            continue
        nom = src_ws.cell(row=r, column=1).value or fund["nom"]
        category = row_to_category.get(r) or ""
        category_fill = row_to_category_fill.get(r)
        for owner, amount in holding_by_owner(src_ws, r, total_col, owner_labels):
            if amount is not None and abs(amount) <= 0.005:
                continue
            selected.append({"isin": isin, "nom": nom, "category": category, "category_fill": category_fill, "owner": owner, "penalite": fund["penalite"]})

    # Regroupement par titulaire : un tableau complètement séparé par titulaire (Monsieur /
    # Madame / société...), plutôt qu'une colonne "Titulaire" au milieu d'un tableau commun —
    # pour que le conseiller voie d'un coup d'œil tout ce qui est détenu par chacun. L'ordre de
    # ces titulaires suit l'ordre de leurs colonnes "Contrat" dans Consolidation (premier
    # titulaire rencontré = premier affiché). Si le fichier n'a qu'un seul titulaire (pas de
    # subdivision par contrat, owner_labels toujours ""), on garde un unique tableau, sans
    # bandeau de titulaire.
    owner_order = []
    by_owner = {}
    for item in selected:
        key = item["owner"] or ""
        if key not in by_owner:
            by_owner[key] = []
            owner_order.append(key)
        by_owner[key].append(item)
    show_owner_headings = len(owner_order) > 1 or (len(owner_order) == 1 and owner_order[0] != "")
    # Harmonisation "M. Prénom NOM" / "Mme Prénom NOM" (uniquement sur cette feuille, jamais sur
    # Consolidation elle-même) à partir du titre de Consolidation, qui énonce explicitement
    # "Monsieur"/"Madame" pour chaque titulaire — jamais deviné. Le bandeau de titulaire reprend
    # le même style que l'en-tête du tableau (pas le style, souvent plus petit, de la ligne où
    # Consolidation affiche "Monsieur"/"Madame" au-dessus de ses colonnes contrat) : c'est un
    # bandeau pleine largeur, pas une étiquette de sous-colonne.
    owner_civilities = extract_owner_civilities(src_ws.cell(row=1, column=1).value) if show_owner_headings else {}

    # Présentation "à la Althos" : bandeau de titre + sous-titre daté (repris tel quel des lignes
    # 1 et 3 de Consolidation — même couleur, même police, même mise en italique), en-tête de
    # tableau dans le même bleu que Consolidation et le titre, bandeaux de catégorie en beige
    # (repris de la couleur des catégories dans Consolidation), fonds sur fond blanc, quadrillage
    # fin en beige. Police commune à toute la feuille (celle de l'en-tête Consolidation, pas un
    # choix arbitraire).
    header_style = copy(src_ws.cell(row=header_row, column=1)._style)  # navy header (theme1)
    header_font = src_ws.cell(row=header_row, column=1).font
    data_font = Font(name=header_font.name, size=header_font.size, bold=False)
    bold_data_font = Font(name=header_font.name, size=header_font.size, bold=True)

    title_style = copy(src_ws.cell(row=1, column=1)._style)
    subtitle_style = copy(src_ws.cell(row=3, column=1)._style)
    # Police de bandeau de catégorie : reprise de la première ligne de catégorie trouvée dans
    # Consolidation (même police en gras). Le FOND, lui, varie par catégorie (voir plus bas) :
    # Consolidation utilise 2 teintes de beige distinctes selon le niveau d'imbrication.
    category_font = copy(src_ws.cell(row=cat_rows[0], column=1).font) if cat_rows else None
    grid_border = build_grid_border(src_ws, header_row)
    white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    ws.sheet_view.showGridLines = False  # comme Consolidation : pas de quadrillage Excel par défaut
    # Vue "aperçu des sauts de page", comme Consolidation : c'est ce qui affiche automatiquement
    # le contour bleu de la zone d'impression en vue Normale.
    ws.sheet_view.view = "pageBreakPreview"

    headers = ["Fonds", "ISIN", "Date d'investissement",
               "Rachat — ordre avant", "Rachat — VL", "Rachat — exécuté",
               "Rachat — publié", "Rachat — cash reçu", "Pénalité de sortie"]
    # 1: titre, 2: (vide), 3: sous-titre daté, 4: (vide), 5: début du tableau — soit directement
    # la ligne d'en-tête (un seul titulaire), soit le 1er bandeau de titulaire suivi de sa propre
    # ligne d'en-tête (plusieurs titulaires).
    HEADER_ROW_OUT = 5
    FIRST_DATA_ROW = HEADER_ROW_OUT + 1

    title_cell = ws.cell(row=1, column=1, value="Calendrier des délais de sortie")
    title_cell._style = copy(title_style)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = src_ws.row_dimensions[1].height or 22.5

    subtitle_cell = ws.cell(row=3, column=1, value=format_date_fr(datetime.date.today()))
    subtitle_cell._style = copy(subtitle_style)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws.row_dimensions[3].height = src_ws.row_dimensions[3].height or 14

    # Ligne d'en-tête ("Fonds | ISIN | ..."), réutilisable : écrite une seule fois si le fichier
    # n'a qu'un seul titulaire, ou répétée juste sous chaque bandeau de titulaire sinon — pour que
    # chaque section du tableau soit lisible seule, sans remonter chercher les intitulés.
    def write_header_row(at_row):
        for i, label in enumerate(headers):
            cell = ws.cell(row=at_row, column=i + 1, value=label)
            cell._style = copy(header_style)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = grid_border
        ws.row_dimensions[at_row].height = 22

    # Bandeau "Si Date de Retrait Exécuté : XX/XX/XXXX" affiché juste sous chaque en-tête de
    # tableau UNIQUEMENT quand le conseiller a choisi de simuler une date de retrait future
    # (sinon rien à signaler, les calculs sont déjà "à la date du jour" par défaut).
    def write_sim_banner(at_row):
        cell = ws.cell(row=at_row, column=1, value=f'="Si Date de Retrait Exécuté : "&TEXT({RD}$1,"dd/mm/yyyy")')
        for c in range(1, len(headers) + 1):
            ws.cell(row=at_row, column=c)._style = copy(header_style)
        ws.merge_cells(start_row=at_row, start_column=1, end_row=at_row, end_column=len(headers))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[at_row].height = 20

    if not show_owner_headings:
        write_header_row(HEADER_ROW_OUT)
        if is_simulating:
            write_sim_banner(FIRST_DATA_ROW)

    PENALTY_COL_WIDTH = 63
    # Facteur 0.7 : la police réelle (Montserrat) est plus large qu'une police système classique
    # et le retour à la ligne se fait sur des mots entiers (jamais exactement à la largeur max),
    # donc le nombre de caractères qui tiennent réellement sur une ligne est nettement inférieur
    # à la largeur de colonne brute — mieux vaut sur-estimer le nombre de lignes que couper le texte.
    CHARS_PER_LINE = round(PENALTY_COL_WIDTH * 0.55)
    widths = [37, 16, 22, 22, 15, 19, 18, 20, PENALTY_COL_WIDTH]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    for name in HELPER_NAMES:
        ws.column_dimensions[helper_col(name)].hidden = True

    if not selected:
        msg = ("Aucun fonds avec calendrier de sortie (rachat) connu n'est actuellement détenu par "
               'ce client (montant total nul, ou fonds hors périmètre "fonds non cotés suivis").')
        ws.cell(row=FIRST_DATA_ROW, column=1, value=msg)
        ws.merge_cells(start_row=FIRST_DATA_ROW, start_column=1, end_row=FIRST_DATA_ROW, end_column=len(headers))
        ws.cell(row=FIRST_DATA_ROW, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[FIRST_DATA_ROW].height = 30
        apply_print_setup(ws, FIRST_DATA_ROW, len(headers))
        return ws, 0, len(fund_rows), None, None

    r = HEADER_ROW_OUT if show_owner_headings else FIRST_DATA_ROW + (1 if is_simulating else 0)
    for owner_idx, owner_key in enumerate(owner_order):
        if show_owner_headings:
            heading = harmonize_owner_label(owner_key, owner_civilities)
            heading_cell = ws.cell(row=r, column=1, value=heading or "Autres titulaires")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c)._style = copy(header_style)
            heading_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            ws.row_dimensions[r].height = 20
            r += 1
            write_header_row(r)
            r += 1
            if is_simulating:
                write_sim_banner(r)
                r += 1

        current_category = object()  # sentinelle : force le 1er bandeau même si catégorie ""
        for item in by_owner[owner_key]:
            if category_font is not None and item["category"] != current_category:
                current_category = item["category"]
                cat_cell = ws.cell(row=r, column=1, value=current_category or "Autres fonds")
                # Fond repris de la VRAIE ligne de catégorie d'origine dans Consolidation (pas un
                # modèle unique emprunté à la 1re catégorie trouvée) : Consolidation utilise 2
                # teintes de beige distinctes selon le niveau d'imbrication.
                band_fill = copy(item["category_fill"]) if item.get("category_fill") is not None else PatternFill(fill_type=None)
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = copy(category_font)
                    cell.fill = copy(band_fill)
                    cell.border = grid_border
                cat_cell.alignment = Alignment(vertical="center")
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
                ws.row_dimensions[r].height = 20
                r += 1

            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = white_fill
                ws.cell(row=r, column=c).border = grid_border

            isin, nom = item["isin"], item["nom"]
            ws.cell(row=r, column=1, value=nom).font = data_font
            ws.cell(row=r, column=2, value=isin).font = data_font
            for col in (1, 2):
                ws.cell(row=r, column=col).alignment = Alignment(vertical="center")

            date_cell = ws.cell(row=r, column=3)
            date_cell.number_format = "dd/mm/yyyy"
            date_cell.font = data_font
            date_cell.alignment = Alignment(horizontal="center", vertical="center")

            b = f'"{isin}"'
            c = f"$C{r}"

            L, M, N = helper_col("has_sortie"), helper_col("next_cutoff_sortie"), helper_col("next_val_sortie")
            Nx, Np, O = helper_col("next_exec_sortie"), helper_col("next_pub_sortie"), helper_col("next_cash_sortie")
            Q = helper_col("months_held")
            R = helper_col("pen_found")
            S = helper_col("kind")
            Tc = helper_col("raw")
            Dv = helper_col("duree_vie")
            U1, V1, W1, X1, Y1, Z1, AA1, AB1, AC1 = (helper_col("max1"), helper_col("rate1"), helper_col("max2"),
                                                      helper_col("rate2"), helper_col("max3"), helper_col("rate3"),
                                                      helper_col("max4"), helper_col("rate4"), helper_col("rate5"))
            AD1 = helper_col("rate_now")

            ws[f"{L}{r}"] = f'=COUNTIFS({cal("A")},{b},{cal("C")},"Rachat")'
            ws[f"{M}{r}"] = (f'=IF({L}{r}=0,"",IFERROR(_xlfn.MINIFS({cal("D")},'
                              f'{cal("A")},{b},{cal("C")},"Rachat",'
                              f'{cal("D")},">="&{RD}$1),""))')
            # VL / exécuté / publié / cash reçu de CETTE échéance précise : on réutilise MINIFS
            # avec une égalité exacte sur la date de cut-off déjà trouvée (au lieu d'une
            # reconstruction de clé texte + MATCH, plus fragile) — même mécanisme que {M}
            # ci-dessus, qui fonctionne de façon fiable. MINIFS ignore les cellules vides : si ce
            # champ précis n'est pas renseigné dans la base pour cette échéance, MINIFS ne trouve
            # aucune valeur numérique et renvoie 0 (jamais une vraie erreur) — sans la
            # vérification "=0" ci-dessous, Excel afficherait ce 0 comme une date, "00/01/1900",
            # au lieu de laisser la cellule vide.
            def minifs_field(col):
                expr = f'_xlfn.MINIFS({cal(col)},{cal("A")},{b},{cal("C")},"Rachat",{cal("D")},{M}{r})'
                return f'=IF({M}{r}="","",IFERROR(IF({expr}=0,"",{expr}),""))'

            ws[f"{N}{r}"] = minifs_field("E")
            ws[f"{Nx}{r}"] = minifs_field("F")
            ws[f"{Np}{r}"] = minifs_field("G")
            ws[f"{O}{r}"] = minifs_field("H")

            ws[f"{Q}{r}"] = f'=IF({c}="","",IF({c}>{RD}$1,"FUTUR",DATEDIF({c},{RD}$1,"m")))'
            ws[f"{R}{r}"] = f'=COUNTIF({pen("A")},{b})'
            ws[f"{S}{r}"] = f'=IF({R}{r}=0,"inconnue",INDEX({pen("C")},MATCH({b},{pen("A")},0)))'
            ws[f"{Tc}{r}"] = f'=IF({R}{r}=0,"",INDEX({pen("M")},MATCH({b},{pen("A")},0)))'
            ws[f"{Dv}{r}"] = f'=IF({R}{r}=0,"",INDEX({pen("N")},MATCH({b},{pen("A")},0)))'
            ws[f"{U1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("D")},MATCH({b},{pen("A")},0)))'
            ws[f"{V1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("E")},MATCH({b},{pen("A")},0)))'
            ws[f"{W1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("F")},MATCH({b},{pen("A")},0)))'
            ws[f"{X1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("G")},MATCH({b},{pen("A")},0)))'
            ws[f"{Y1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("H")},MATCH({b},{pen("A")},0)))'
            ws[f"{Z1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("I")},MATCH({b},{pen("A")},0)))'
            ws[f"{AA1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("J")},MATCH({b},{pen("A")},0)))'
            ws[f"{AB1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("K")},MATCH({b},{pen("A")},0)))'
            ws[f"{AC1}{r}"] = f'=IF({R}{r}=0,0,INDEX({pen("L")},MATCH({b},{pen("A")},0)))'
            ws[f"{AD1}{r}"] = (f'=IF(OR({c}="",{Q}{r}="FUTUR"),"",_xlfn.IFS({Q}{r}<{U1}{r},{V1}{r},{Q}{r}<{W1}{r},{X1}{r},'
                                f'{Q}{r}<{Y1}{r},{Z1}{r},{Q}{r}<{AA1}{r},{AB1}{r},TRUE(),{AC1}{r}))')

            # Colonnes visibles D..H : valeurs reprises telles quelles de la base (une par champ).
            for col_letter, helper in (("D", M), ("E", N), ("F", Nx), ("G", Np), ("H", O)):
                cell = ws[f"{col_letter}{r}"]
                cell.value = f"={helper}{r}"
                cell.number_format = "dd/mm/yyyy"
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Pénalité de sortie : vide dès qu'il n'y a rien d'actionnable à signaler — pas de
            # pénalité prévue pour ce fonds (kind="aucune"), aucune pénalité renseignée dans la
            # base (kind="inconnue", traité comme "pas de pénalité"), ou délai de pénalité
            # dépassé. Un message n'apparaît que dans les cas où le conseiller doit agir : fonds
            # fermé (sortie impossible hors cas exceptionnel), règle ambiguë à vérifier à la
            # main, ou pénalité active à signaler au client.
            ws[f"I{r}"] = (
                f'=_xlfn.IFS('
                f'{S}{r}="ferme","Fonds fermé : aucun rachat possible."&IF({Dv}{r}<>"",CHAR(10)&{Dv}{r},""),'
                f'{S}{r}="manuel","À VÉRIFIER MANUELLEMENT : "&{Tc}{r},'
                f'{S}{r}="aucune","",'
                f'{S}{r}="inconnue","",'
                f'{c}="","Saisir une date d\'investissement pour statuer sur la pénalité.",'
                f'{Q}{r}="FUTUR","Date d\'investissement postérieure à aujourd\'hui — vérifier la saisie.",'
                f'{AD1}{r}>0,"Pénalité de sortie : "&{Tc}{r}&CHAR(10)&"Durée de détention actuelle : "&{Q}{r}&" mois.",'
                f'TRUE(),"")'
            )
            pen_cell = ws[f"I{r}"]
            # Une formule Excel ne peut jamais renvoyer un texte à mise en forme mixte (gras
            # partiel) : tout le message (pénalité ou fermeture) est mis en gras en bloc, comme
            # la colonne "Rachat — cash reçu".
            pen_cell.font = bold_data_font
            pen_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            # Hauteur de 60pt par défaut (confortable pour l'immense majorité des messages, 2-3
            # lignes) ; augmentée seulement si le texte réel de la base pour ce fonds précis est
            # inhabituellement long (ça arrive), pour ne jamais le couper.
            pen_info = item.get("penalite") or {"kind": "inconnue", "raw": None}
            lines = estimate_penalty_lines(pen_info.get("kind"), len(pen_info.get("raw") or ""), len(pen_info.get("dureeVie") or ""), CHARS_PER_LINE)
            ws.row_dimensions[r].height = max(DEFAULT_PENALTY_ROW_HEIGHT, height_for_lines(lines, data_font.size))
            r += 1

        # 1 ligne vide (non bordée) entre 2 titulaires, pour que 2 tableaux distincts ne
        # paraissent pas collés l'un à l'autre.
        if show_owner_headings and owner_idx < len(owner_order) - 1:
            r += 1

    apply_print_setup(ws, r - 1, len(headers))
    return ws, len(selected), len(fund_rows), FIRST_DATA_ROW, r - 1


def apply_print_setup(ws, last_row, last_visible_col):
    """Zone d'impression = tout le tableau, mise à l'échelle sur une page en largeur, paysage —
    comme sur la feuille Consolidation. Marge de 1 ligne/colonne entre le contenu et le contour
    de la zone d'impression. Pour les colonnes, "le contenu" inclut les colonnes d'aide masquées
    (jusqu'à HELPER_FIRST_COL + len(HELPER_NAMES) - 1), pas seulement les colonnes visibles :
    sinon le contour passerait au milieu de colonnes masquées au lieu de se terminer 1 colonne
    après elles."""
    last_helper_col = HELPER_FIRST_COL + len(HELPER_NAMES) - 1
    last_col = max(last_visible_col, last_helper_col)
    ws.print_area = f"A1:{get_column_letter(last_col + 1)}{last_row + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_consolidation_columns(src_ws, header_row, total_col, exit_sheet_name, first_data_row, last_data_row, funds_by_isin):
    """Ajoute 2 colonnes directement sur Consolidation ("Rachat — cash reçu" et "Pénalité de
    sortie"), une par fonds réellement détenu, en formule vers la feuille "Calendrier de sortie"
    déjà construite : même information dans les 2 pages, une seule source de vérité (pas de
    calcul dupliqué)."""
    if not last_data_row or last_data_row < first_data_row:
        return  # aucun fonds retenu, rien à référencer

    # "Mouvements en cours" (si présente) est toujours la toute dernière colonne du tableau
    # Consolidation, mais n'en fait pas vraiment partie : les 2 nouvelles colonnes doivent être
    # collées juste après la dernière colonne réellement utile (ex. "Dernière Valeur
    # Liquidative"), et "Mouvements en cours" repoussée de 2 colonnes vides après elles.
    mov_col = find_column_by_header_text(src_ws, header_row, "Mouvements en cours")
    if mov_col:
        core_last = mov_col - 1
        while core_last >= 1:
            v = src_ws.cell(row=header_row, column=core_last).value
            if v is not None and str(v).strip() != "":
                break
            core_last -= 1
        cash_col = core_last + 1
        pen_col = core_last + 2
        mov_end = merged_block_end_column(src_ws, header_row, mov_col)
        desired_mov_start = core_last + 5  # cash_col, pen_col, 2 colonnes vides, puis Mouvements en cours
        if desired_mov_start > mov_col:
            shift_column_block(src_ws, mov_col, mov_end, desired_mov_start)
    else:
        last_col = find_last_header_column(src_ws, header_row)
        cash_col = next_safe_column(src_ws, header_row, last_col + 1)
        pen_col = next_safe_column(src_ws, header_row, cash_col + 1)

    header_style = copy(src_ws.cell(row=header_row, column=1)._style)
    header_font = src_ws.cell(row=header_row, column=1).font
    data_font = Font(name=header_font.name, size=header_font.size, bold=False)
    bold_data_font = Font(name=header_font.name, size=header_font.size, bold=True)
    two_row_header = has_two_row_header(src_ws, header_row, 1)
    grid_border = build_grid_border(src_ws, header_row)
    PENALTY_COL_WIDTH = 63  # même largeur que sur "Calendrier de sortie", pour un rendu cohérent
    CHARS_PER_LINE = round(PENALTY_COL_WIDTH * 0.55)  # cf. commentaire dans build_exit_sheet

    for col, label, width in ((cash_col, "Rachat — cash reçu", 16), (pen_col, "Pénalité de sortie", PENALTY_COL_WIDTH)):
        cell = src_ws.cell(row=header_row, column=col, value=label)
        cell._style = copy(header_style)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = grid_border
        col_letter = get_column_letter(col)
        src_ws.column_dimensions[col_letter].width = width
        # Toujours visible : la colonne peut hériter un état masqué du fichier d'origine si sa
        # lettre correspondait déjà à une colonne cachée (ex. une colonne d'aide de l'ancien
        # "Mouvements en cours" déplacé) sans qu'on l'ait explicitement remise à zéro.
        src_ws.column_dimensions[col_letter].hidden = False
        if two_row_header:
            bottom_cell = src_ws.cell(row=header_row + 1, column=col)
            bottom_cell._style = copy(header_style)
            bottom_cell.border = grid_border
            src_ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row + 1, end_column=col)

    isin_range = f"'{exit_sheet_name}'!$B${first_data_row}:$B${last_data_row}"
    cash_range = f"'{exit_sheet_name}'!$H${first_data_row}:$H${last_data_row}"
    pen_range = f"'{exit_sheet_name}'!$I${first_data_row}:$I${last_data_row}"

    # Style de bandeau de catégorie (fond beige) : Consolidation utilise 2 teintes de beige
    # distinctes selon le niveau d'imbrication (ex. catégorie principale en beige foncé,
    # sous-catégorie en beige clair) — on reprend donc le fond de CHAQUE ligne de catégorie
    # individuellement, pas un seul modèle emprunté à la 1re catégorie trouvée.
    category_rows, fund_rows, _, _ = classify_rows(src_ws, header_row, total_col)

    # cash_col/pen_col occupent une position de colonne qui existait déjà dans le fichier
    # d'origine (juste après la dernière colonne utile, ou l'ancien emplacement de "Mouvements en
    # cours") : ses cellules peuvent donc porter un fond/quadrillage hérité du fichier client (ex.
    # un second petit tableau récapitulatif de répartition, hors du tableau principal, plus bas
    # sur la feuille). On repart d'une ardoise vierge sur toute la hauteur avant de ne redessiner
    # que les lignes catégorie/fonds du VRAI tableau, pour ne jamais laisser un bloc beige
    # résiduel sans quadrillage.
    no_fill = PatternFill(fill_type=None)
    no_border = Border()
    clear_from_row = header_row + (2 if two_row_header else 1)  # ne jamais effacer la 2e ligne d'un en-tête fusionné sur 2 lignes
    for r in range(clear_from_row, src_ws.max_row + 1):
        for col in (cash_col, pen_col):
            cell = src_ws.cell(row=r, column=col)
            cell.fill = copy(no_fill)
            cell.border = copy(no_border)

    # Le quadrillage doit courir sans interruption sur TOUTE la hauteur du VRAI tableau
    # (catégories ET fonds, détenus ou non) — sinon chaque fonds non détenu par ce client laisse
    # un "trou" dans les 2 nouvelles colonnes, puisque Consolidation liste l'univers complet des
    # fonds, pas seulement ceux de ce client.
    for r in category_rows:
        row_fill = copy(src_ws.cell(row=r, column=1).fill)
        for col in (cash_col, pen_col):
            cell = src_ws.cell(row=r, column=col)
            cell.fill = copy(row_fill)
            cell.border = grid_border
    for r in fund_rows:
        for col in (cash_col, pen_col):
            src_ws.cell(row=r, column=col).border = grid_border

    for r in fund_rows:
        isin_raw = src_ws.cell(row=r, column=2).value
        isin = isin_raw.strip() if isinstance(isin_raw, str) else isin_raw
        if not isin:
            continue
        amount = holding_amount(src_ws, r, total_col)
        is_held = amount is None or abs(amount) > 0.005
        if not is_held:
            continue

        b = f'"{isin}"'
        cash_cell = src_ws.cell(row=r, column=cash_col, value=f"=IFERROR(INDEX({cash_range},MATCH({b},{isin_range},0)),\"\")")
        cash_cell.number_format = "dd/mm/yyyy"
        cash_cell.font = bold_data_font
        cash_cell.alignment = Alignment(horizontal="center", vertical="center")

        pen_cell = src_ws.cell(row=r, column=pen_col, value=f"=IFERROR(INDEX({pen_range},MATCH({b},{isin_range},0)),\"\")")
        # Gras en bloc (cf. commentaire dans build_exit_sheet).
        pen_cell.font = bold_data_font
        pen_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Hauteur de 60pt par défaut, augmentée seulement si le texte réel de la base pour ce
        # fonds précis est inhabituellement long — remplace toute hauteur figée héritée du
        # fichier d'origine du conseiller (potentiellement trop courte pour ce nouveau texte).
        fund = funds_by_isin.get(isin)
        pen_info = (fund or {}).get("penalite") or {"kind": "inconnue", "raw": None}
        lines = estimate_penalty_lines(pen_info.get("kind"), len(pen_info.get("raw") or ""), len(pen_info.get("dureeVie") or ""), CHARS_PER_LINE)
        src_ws.row_dimensions[r].height = max(DEFAULT_PENALTY_ROW_HEIGHT, height_for_lines(lines, data_font.size))

    extend_merge_right(src_ws, 1, pen_col)
    extend_print_area(src_ws, pen_col)


def main():
    src_path = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "source" / "ConsolidationTemplateAlthosAI_V5.xlsx"
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "output" / "ConsolidationTemplateAlthosAI_V5_avec_calendrier.xlsx"
    # Date de retrait simulée (optionnelle), au format AAAA-MM-JJ : si fournie, TOUT le tableau
    # (échéances de rachat, durée de détention, pénalité) est calculé comme si on s'y trouvait
    # déjà, plutôt qu'à la date du jour.
    ref_date_iso = sys.argv[3] if len(sys.argv) > 3 else None
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb_cal = openpyxl.load_workbook(CAL_FILE, data_only=True)
    funds = load_suivi(wb_cal)
    funds = merge_extra_from_calendriers_par_fonds(wb_cal, funds)
    funds = attach_duree_vie(wb_cal, funds)
    calendar = load_calendrier(wb_cal)
    calendar = merge_calendriers(calendar, load_calendrier_par_fonds(wb_cal))
    for f in funds.values():
        f["hasCalendar"] = bool(f["isin"] and f["isin"] in calendar)
    funds_by_isin = {f["isin"]: f for f in funds.values() if f["isin"]}

    wb = openpyxl.load_workbook(src_path, data_only=False)
    src_ws = wb["Consolidation"]
    # Force Excel à recalculer TOUTES les formules à l'ouverture (le fichier n'a pas de
    # calcChain.xml puisqu'openpyxl n'exécute pas les formules lui-même) : évite les cellules
    # calculées qui restent vides tant que l'utilisateur n'a pas forcé un recalcul manuel.
    wb.calculation.fullCalcOnLoad = True

    ws_cal = write_bdd_calendrier(wb, calendar)
    ws_pen = write_bdd_penalites(wb, funds)
    exit_ws, selected_count, fund_rows_count, first_data_row, last_data_row = build_exit_sheet(
        wb, src_ws, calendar, ws_cal.max_row, ws_pen.max_row, funds_by_isin, ref_date_iso)

    # Complète aussi Consolidation elle-même avec 2 colonnes (cash reçu / pénalité de sortie),
    # en formule vers "Calendrier de sortie" — même information, une seule source de vérité.
    header_row = find_header_row(src_ws)
    total_col = find_total_column(src_ws, header_row)
    add_consolidation_columns(src_ws, header_row, total_col, exit_ws.title, first_data_row, last_data_row, funds_by_isin)

    wb.active = wb.sheetnames.index("Calendrier de sortie")
    wb.save(out_path)
    print(f"OK -> {out_path} ({selected_count} fonds retenus sur {fund_rows_count} lignes analysées)")


if __name__ == "__main__":
    main()
